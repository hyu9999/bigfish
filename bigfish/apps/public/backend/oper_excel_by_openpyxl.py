import os

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "bigfish.settings.base")
django.setup()
from openpyxl import Workbook, load_workbook
from bigfish.apps.public.models import Public, AppTable


class OperateExcel:
    """
    excel操作类

    :params:

        {
            "app": "visualbackend",
            "model": "StageScore",
            "query_filter" :{"term":1},
            "file_path":"a.xlsx",
            "sheet_name":"sheet1",
            "header":["school_name", "term_id", "school_term_id"]
        }
    """
    sheet_name = 'sheet1'
    wb = None

    def __init__(self, params):
        self.file_path = params.get('file_path')
        # assert os.path.isfile(self.file_path), "[{}]传入路径异常!".format(self.file_path)
        eval_str = "from bigfish.apps.{}.models import {}".format(params.get('app'), params.get('model'))
        exec(eval_str)
        self.model_obj = eval(params.get('model'))
        self.query_filter = params.get('query_filter')
        self.header = params.get('header')
        self.sheet_name = params.get('sheet_name') or self.sheet_name

    def get_workbook(self, flag):
        """
        获取workbook对象

        :param flag:
        :return:
        """
        wb = None
        if flag == 'e':
            wb = Workbook()
        elif flag == 'i':
            wb = load_workbook(self.file_path)
        self.wb = wb
        return wb

    def get_worksheet(self, flag=None, sheet_name=None):
        """
        返回Worksheet对象（openpyxl.worksheet.worksheet.Worksheet）

        :param flag:
        :param sheet_name:
        :return:
        """
        wb = self.get_workbook(flag)
        self.sheet_name = sheet_name or self.sheet_name
        if not self.sheet_name:
            ws = wb.active
        else:
            if self.sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        return ws

    def wb_save(self):
        """
        workbook对象保存到文件

        :return:
        """
        self.wb.save(self.file_path)

    def get_header(self):
        """
        获取头部信息

        :return:
        """
        return self.header

    def save_to_file(self, header=None, file_path=None):
        """
        数据存入文件（xlsx/xls）

        :return:
        """
        ws = self.get_worksheet('e', self.sheet_name)  # 获取worksheet对象
        self.header = header or self.header
        if self.header:
            ws.append(header)
            queryset = self.model_obj.objects.filter(**self.query_filter).values_list(*header)
        else:
            queryset = self.model_obj.objects.filter(**self.query_filter).values_list()
        for r in queryset:
            ws.append(r)
        self.file_path = file_path or self.file_path
        self.wb_save()
        return self.file_path

    def import_data(self, header=None):
        """
        导入数据库

        :param header:
        :return:
        """
        ws = self.get_worksheet('i', self.sheet_name)
        print(self.sheet_name)
        self.header = header or self.header
        if self.header:
            index = 2
        else:
            index = 1
        sql_list = []
        for row in range(index, ws.max_row + 1):
            column_val_list = [ws.cell(row=row, column=x).value for x in range(1, ws.max_column + 1)]
            param_dict = dict(zip(self.header, column_val_list))
            print(param_dict)
            tmp_data = self.model_obj(**param_dict)
            sql_list.append(tmp_data)
        ret_queryset = self.model_obj.objects.bulk_create(sql_list)
        return ret_queryset


def test_app_table():
    table_name = "AppTable"
    table_verbose_name = AppTable._meta.verbose_name_plural
    print("table_name=", table_name)
    print("table_verbose_name=", table_verbose_name)
    app_name = Public._meta.app_label
    app_verbose_name = Public._meta.app_config.verbose_name
    print("app_name=", app_name)
    print("app_verbose_name=", app_verbose_name)
    # for item in Public._meta.apps.all_models.items():
    #     print(
    #         item
    #     )
    #     print("app name===========", item[0])
    #     print("app value===========", item[0]._meta.verbose_name)
    #     for j in item[1].items():
    #         print("     table name==========={}".format(j[0]))
    #         if j[1]:
    #             print("     table value===========", j[1]._meta.verbose_name_plural)
    # for item in Public._meta.app_label:
    #     print(
    #         item
    #     )
    #     print("app name===========", item[0])
    #     print("app value===========", item[0]._meta.verbose_name)
    #     for j in item[1].items():
    #         print("     table name==========={}".format(j[0]))
    #         if j[1]:
    #             print("     table value===========", j[1]._meta.verbose_name_plural)



if __name__ == '__main__':
    # params = {
    #     "app": "visualbackend",
    #     "model": "StageScore",
    #     "query_filter": {"term_id": 1},
    #     "file_path": "a.xlsx",
    #     "sheet_name": "sheet1",
    #     "header": ["school_name", "term_id", "school_term_id"]
    # }
    # # ie = OperateExcel(params)
    # # ie.save_to_file(['school_name', 'term_id', 'school_term_id'])
    # # ie.import_data(['school_name', 'term_id', 'school_term_id'])
    # str = "for i in range(0,10): print(i)"
    # c = compile(str, '', 'exec')
    # print(exec(str))
    test_app_table()
