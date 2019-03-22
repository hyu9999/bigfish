import logging

from django.conf import settings
from django.db import transaction
from django.db.models import Sum
from openpyxl import load_workbook
from rest_framework.decorators import list_route
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.response import Response

from bigfish.apps.reports.models import ExaminationReport, RatingReport, PracticalCourseRecord
from bigfish.apps.reports.serializers import ExaminationReportSerializer, RatingReportSerializer, \
    PracticalCourseRecordSerializer
from bigfish.apps.reports.tools import create_practical_course_record
from bigfish.base import viewsets, status
from bigfish.base.response import BFValidationError
from bigfish.base.ret_msg import rsp_msg_200
from bigfish.utils.functions import generate_fields, \
    get_field_name

logger = logging.getLogger('django')


class ExaminationReportViewSet(viewsets.ModelViewSet):
    """
    retrieve:
        Return a ExaminationReport instance.
    list:
        Return all ExaminationReport, ordered by most recently joined.
    create:
        Create a new ExaminationReport.
    delete:
        Remove an existing ExaminationReport.
    partial_update:
        Update one or more fields on an existing ExaminationReport.
    update:
        Update a ExaminationReport.
    """
    queryset = ExaminationReport.objects.all()
    serializer_class = ExaminationReportSerializer
    parser_classes = (JSONParser, MultiPartParser)
    filter_fields = generate_fields(ExaminationReport)

    @list_route(methods=["post"])
    def import_examination_report(self, request):
        """
        导入测试成绩（以班为单位导入）

        :param request:\n
            {
                "file_url":"/media/成绩导入表.xlsx"
            }
        :return:\n
            {
                "detail": "success",
                "code": 200,
                "results": {
                    "score": 5
                },
                "error_msg":"xxx"
            }
        """

        file_url = request.data.get("file_url", "")
        file_path = settings.MEDIA_ROOT.replace("media", "") + file_url

        wb = load_workbook(filename=file_path)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        verbose_name_list = []
        filter_dict = {}

        for column in ws.columns:
            column_value = column[0].value
            if column_value == '学校编码':
                distinct_value = list(set([x.value for x in column[1:] if x.value]))
                if distinct_value.__len__() != 1:
                    raise ValidationError("学校编码不一致！")
                else:
                    filter_dict['school_code'] = distinct_value[0]
            if column_value == '班级':
                distinct_value = list(set([x.value for x in column[1:] if x.value]))
                if distinct_value.__len__() != 1:
                    raise ValidationError("班级不一致！")
                else:
                    filter_dict['grade'] = distinct_value[0]
            if column_value == '学期':
                distinct_value = list(set([x.value for x in column[1:] if x.value]))
                if distinct_value.__len__() != 1:
                    raise ValidationError("学期不一致！")
                else:
                    filter_dict['term'] = distinct_value[0]
            if column_value == '学年':
                distinct_value = list(set([x.value for x in column[1:] if x.value]))
                if distinct_value.__len__() != 1:
                    raise ValidationError("学年不一致！")
                else:
                    filter_dict['academic_year'] = distinct_value[0]
            verbose_name_list.append(column_value)

        query_field_list = get_field_name(ExaminationReport, *verbose_name_list)
        if len(query_field_list) != len(verbose_name_list):
            raise ValidationError('请检查列名是否有误')
        rating_report_sql_list = []
        with transaction.atomic():
            try:
                list_queryset = import_excel_tool(ExaminationReport, query_field_list, file_url.replace("media", ""))
            except Exception as e:
                logger.debug(e)
                raise ValidationError("写入数据异常！[{}]".format(str(e)))
            queryset = ExaminationReport.objects.filter(**filter_dict)
            user_list = list(set([(x['username'], x['realname']) for x in list_queryset]))

            for obj in user_list:
                username = obj[0]
                tmp_dict = {'username': obj[0], 'realname': obj[1]}
                tmp_dict.update(filter_dict)
                all_cnt = queryset.filter(username=username).count()
                if all_cnt == 0:
                    continue
                examine_list = queryset.filter(username=username, exam_type__startswith='普通测试')

                examine_cnt = examine_list.count()
                weight_power = 1
                if examine_cnt != 0 and all_cnt == examine_cnt:
                    weight_power = float(examine_cnt / (3 * all_cnt))

                try:
                    avg_score = examine_list.aggregate(Sum('score'))['score__sum'] / (3 * all_cnt)
                except Exception as e:
                    avg_score = 0
                    logger.debug(e)
                other_list = \
                    queryset.filter(username=username).exclude(exam_type__startswith='普通测试').aggregate(Sum('score'))[
                        'score__sum']
                try:
                    avg_score += other_list * (1 - examine_cnt / (3 * all_cnt)) / (all_cnt - examine_cnt)
                except Exception as e:
                    logger.debug(e)
                sql = RatingReport.objects.get_or_create(**tmp_dict)[0]
                # 除以权重
                final_score = round(avg_score / weight_power)
                if final_score >= 60:
                    sql.level = 1
                else:
                    sql.level = 2
                sql.save()
                rating_report_sql_list.append(RatingReportSerializer(sql).data)
        return Response(rsp_msg_200(rating_report_sql_list), status=status.HTTP_200_OK)

    @list_route(methods=['POST'])
    def sum_examine_data(self, request):
        """
         汇总测试数据

        :param request:\n
            {
                "school_code": 2700040,
                "grade": "小学四年级",
                "term": 1,
                "academic_year": "2018"
            }
        :return:\n
            {
                "detail": "success",
                "code": 200,
                "results": {
                    "score": 5
                },
                "error_msg":"xxx"
            }
        """
        queryset = self.filter_queryset(self.get_queryset())
        user_list = queryset.order_by('school_code', 'grade', 'username', 'realname', 'term', 'academic_year').values(
            'school_code', 'grade', 'username', 'realname', 'term', 'academic_year').distinct()
        ret_list = []
        for obj in user_list:
            username = obj['username']
            tmp_dict = {'username': username}
            all_cnt = queryset.filter(username=username).count()
            examine_list = queryset.filter(username=username, exam_type__startswith='普通测试')
            examine_cnt = examine_list.count()
            try:
                avg_score = examine_list.aggregate(Sum('score'))['score__sum'] / (3 * all_cnt)
            except Exception as e:
                avg_score = 0
                logger.debug(e)
            other_list = \
                queryset.filter(username=username).exclude(exam_type__startswith='普通测试').aggregate(Sum('score'))[
                    'score__sum']
            try:
                avg_score += other_list * (1 - examine_cnt / (3 * all_cnt)) / 2
            except Exception as e:
                logger.debug(e)

            tmp_dict['avg_score'] = round(avg_score)
            # level = 2
            # if avg_score >= 60:
            #     level = 1
            # RatingReport.objects.update_or_create(defaults={'level': level}, **obj)
            # update userprofile
            # try:
            #     user_profile = UserProfile.objects.get(user__username=obj['username'])
            # except Exception as e:
            #     logger.debug(e)
            # else:
            #     user_profile.level = level
            #     user_profile.save()
            ret_list.append(tmp_dict)
        return Response(rsp_msg_200(ret_list), status=status.HTTP_200_OK)


def import_excel_tool(model_obj, query_field_list, url_path):
    filename = settings.MEDIA_ROOT + url_path
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    lists = []
    for row in range(2, ws.max_row + 1):
        r = {}
        for col in range(1, len(query_field_list) + 1):
            key = query_field_list[col - 1]
            r[key] = ws.cell(row=row, column=col).value
            if key == 'score':
                r[key] = r[key] * 100
        lists.append(r)
    sql_list = []
    for cell in lists:
        data_dict = {}
        for header in query_field_list:
            data_dict[header] = str(cell[header])
        defaults = {'score': cell['score']}
        kwargs = {}
        kwargs.update(cell)
        kwargs.pop('score', None)
        value_list = [v for v in data_dict.values() if v]
        if value_list:
            sql_list.append(cell)
            model_obj.objects.update_or_create(defaults=defaults, **kwargs)
        else:
            break
    return sql_list


class RatingReportViewSet(viewsets.ModelViewSet):
    """
    retrieve:
        Return a RatingReport instance.
    list:
        Return all RatingReport, ordered by most recently joined.
    create:
        Create a new RatingReport.
    delete:
        Remove an existing RatingReport.
    partial_update:
        Update one or more fields on an existing RatingReport.
    update:
        Update a RatingReport.
    """
    queryset = RatingReport.objects.all()
    serializer_class = RatingReportSerializer
    filter_fields = generate_fields(RatingReport)

    @list_route(methods=["post"])
    def import_student_level(self, request):
        """
        导入学生级别

        :param request:\n
            {
                "file_url":"/media/student_level.xlsx"
            }
        :return:\n
            {
              "result": "success",
              "code": 200,
              "detail": [
                {
                  "id": 1145,
                  "school_code": "2900046",
                  "grade": "小学三年级1班",
                  "term": 2,
                  "academic_year": 2017,
                  "username": "10460001",
                  "realname": "杨丕进",
                  "score": 74,
                  "level": 1,
                  "add_time": "2018-02-26T14:32:51.033401"
                },
                {
                  "id": 1189,
                  "school_code": "2900046",
                  "grade": "小学三年级1班",
                  "term": 2,
                  "academic_year": 2017,
                  "username": "10460045",
                  "realname": "李喆",
                  "score": 34,
                  "level": 2,
                  "add_time": "2018-02-26T14:32:51.372401"
                }
              ]
            }
        """

        file_url = request.data.get("file_url", "")
        file_path = settings.MEDIA_ROOT.replace("media", "") + file_url
        wb = load_workbook(filename=file_path)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        query_field_list = []

        for column in ws.columns:
            column_value = column[0].value
            query_field_list.append(column_value)
        ret_data = []
        with transaction.atomic():
            lists = []
            for row in range(2, ws.max_row + 1):
                r = {}
                for col in range(1, len(query_field_list) + 1):
                    key = query_field_list[col - 1]
                    r[key] = ws.cell(row=row, column=col).value
                lists.append(r)
            for cell in lists:
                kwargs = {}
                default = {}
                for header in query_field_list[:6]:
                    kwargs[header] = str(cell[header])
                for header in query_field_list[6:]:
                    default[header] = str(cell[header])
                    if header == "level":
                        if default[header] == '好':
                            default[header] = 1
                        elif default[header] == '差':
                            default[header] = 2
                        else:
                            continue
                    else:
                        continue
                value_list = [v for v in kwargs.values() if v and v != 'None']
                if value_list:
                    try:
                        obj, flag = RatingReport.objects.update_or_create(defaults=default, **kwargs)
                    except Exception as e:
                        logger.debug(e)
                        continue
                    # try:
                    #     user_profile = UserProfile.objects.get(user__username=kwargs['username'])
                    # except Exception as e:
                    #     logger.debug(e)
                    # else:
                    #     user_profile.level = default['level']
                    #     user_profile.save()
                    ret_data.append(RatingReportSerializer(obj).data)
                else:
                    continue
        data = {"result": "success", "code": 200, "detail": ret_data}
        return Response(rsp_msg_200(data), status=status.HTTP_200_OK)


class PracticalCourseRecordViewSet(viewsets.ModelViewSet):
    """
    retrieve:
        Return a PracticalCourseRecord instance.
    list:
        Return all PracticalCourseRecord, ordered by most recently joined.
    create:
        Create a new PracticalCourseRecord.
    delete:
        Remove an existing PracticalCourseRecord.
    partial_update:
        Update one or more fields on an existing PracticalCourseRecord.
    update:
        Update a PracticalCourseRecord.

    """
    queryset = PracticalCourseRecord.objects.all()
    serializer_class = PracticalCourseRecordSerializer
    filter_fields = ('id', 'klass', 'teacher', 'subject', 'schedule', 'teaching_date')

    @list_route(methods=['POST'])
    def auto_generate_practical_course(self, request):
        """
        自定生成实际教学记录（根据教师课程）

        :param request: \n
            {
            "date_range":["2018-03-01", "2018-06-23"],
            "klass_list":[1,3]
            }
        :return:
        """
        time_range = request.data.get('date_range', [])
        if len(time_range) == 0:
            raise BFValidationError("请传入正确的日期！")
        klass_list = request.data.get("klass_list", [])
        if len(klass_list) == 0:
            ret_data = create_practical_course_record(time_range)
        else:
            ret_data = create_practical_course_record(time_range, klass_list=klass_list)
        data = {"result": "success", "code": 200, "detail": ret_data}
        return Response(rsp_msg_200(data), status=status.HTTP_200_OK)
