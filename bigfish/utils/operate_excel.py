import os

import datetime
from math import modf, ceil

import xlwt
from django.conf import settings
from django.db import models
from django.db.models import Field
from openpyxl import load_workbook
from openpyxl.compat import range
from rest_framework.exceptions import ValidationError


def export_excel(val_list, field_list, filename="data.xlsx", sheet_name="sheet1"):
    file_path = os.path.join(settings.MEDIA_ROOT, filename)
    # new一个文件
    wb = xlwt.Workbook(encoding='utf-8')
    list_len = len(val_list)
    sheet_num = ceil(list_len / 65535) + 1
    for item in range(1, sheet_num):
        # new一个sheet
        sheet = wb.add_sheet("sheet{}".format(item))
        # 维护一些样式， style_heading, style_body, style_red, style_green
        style_heading, style_body = excel_download_style()

        # 写标题栏
        for index, val in enumerate(field_list.values()):
            sheet.write(0, index, val, style_heading)

        # 写数据
        row = 1
        start_index = 0 + 65535 * (item - 1)
        end_index = 65535 * item
        for query_list in val_list[start_index:end_index]:
            for index, key in enumerate(field_list.keys()):
                data_value = query_list[index]
                if isinstance(data_value, datetime.datetime) or isinstance(data_value, datetime.date) or isinstance(
                        data_value, datetime.time):
                    data_value = data_value.__str__()
                sheet.write(row, index, data_value, style_body)
            row = row + 1
    wb.save(file_path)


def excel_download_style():
    # 维护一些样式， style_heading, style_body, style_red, style_green
    style_heading = xlwt.easyxf("""
            font:
                name Arial,
                colour_index white,
                bold on,
                height 0xB0;
            align:
                wrap off,
                vert center,
                horiz center;
            pattern:
                pattern solid,
                fore-colour 0x19;
            borders:
                left THIN,
                right THIN,
                top THIN,
                bottom THIN;
            """
                                )

    style_body = xlwt.easyxf("""
                font:
                    name Arial,
                    bold off,
                    height 0XA0;
                align:
                    wrap on,
                    vert center,
                    horiz left;
                borders:
                    left THIN,
                    right THIN,
                    top THIN,
                    bottom THIN;
                """
                             )
    # style_green = xlwt.easyxf(" pattern: pattern solid,fore-colour 0x11;")
    # style_red = xlwt.easyxf(" pattern: pattern solid,fore-colour 0x0A;")
    # fmts = [
    #     'M/D/YY',
    #     'D-MMM-YY',
    #     'D-MMM',
    #     'MMM-YY',
    #     'h:mm AM/PM',
    #     'h:mm:ss AM/PM',
    #     'h:mm',
    #     'h:mm:ss',
    #     'M/D/YY h:mm',
    #     'mm:ss',
    #     '[h]:mm:ss',
    #     'mm:ss.0',
    # ]
    # style_body.num_format_str = fmts[0]
    return style_heading, style_body


def import_excel(model_obj, query_field_list, url_path):
    filename = settings.MEDIA_ROOT + url_path
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    lists = []
    for row in range(2, ws.max_row + 1):
        r = {}
        for col in range(1, len(query_field_list) + 1):
            key = query_field_list[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sql_list = []
    for cell in lists:
        data_dict = {}
        for header in query_field_list:
            data_dict[header] = str(cell[header])
        value_list = [v for v in data_dict.values() if v]
        if value_list:
            try:
                param_dict = format_param(model_obj, data_dict)
            except Exception as e:
                raise ValidationError("格式化数据异常[{}::{} ]".format(e, value_list))
            sql = model_obj(**param_dict)
            sql_list.append(sql)
        else:
            break
    ret_queryset = model_obj.objects.bulk_create(sql_list)
    return ret_queryset


def format_param(model, data_dict):
    ret_data = {}
    for field in model._meta.get_fields():
        if field.name not in data_dict.keys():
            continue
        if isinstance(field, Field):
            if isinstance(field, models.ForeignKey):
                fk_str = str(field.related_model).split("'")[1]
                try:
                    eval_str = "from {} import {}".format(fk_str[:fk_str.rfind(".")], fk_str[fk_str.rfind(".") + 1:])
                    exec(eval_str)
                except Exception as e:
                    raise ValidationError("数据异常[{}::{}:{} ]".format(e, fk_str, fk_str))
                id_val = int(data_dict[field.name])
                ret_data[field.name] = field.related_model.objects.get(id=id_val)
            else:
                ret_data[field.name] = data_dict[field.name]
    return ret_data
