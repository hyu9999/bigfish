from io import StringIO

import xlwt
from django.conf import settings
from django.forms import model_to_dict
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.compat import range


def import_user(obj, headers=[]):
    wb = load_workbook(filename=settings.MEDIA_ROOT + '/a.xlsx')
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    lists = []
    for row in range(2, ws.max_row + 1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sql_list = []
    for cell in lists:
        data_dict = {}
        for header in headers:
            data_dict[header] = cell[header]
        value_list = [data_dict[x] for x in data_dict if data_dict[x] != None]
        if value_list:
            sql = obj(**data_dict)
            sql_list.append(sql)
        else:
            break
    obj.objects.bulk_create(sql_list)


def export_excel(request, head_list, queryset, filename="data.xls", sheet_name="sheet1"):
    # new一个文件
    wb = xlwt.Workbook(encoding='utf-8')
    # new一个sheet
    sheet = wb.add_sheet(sheet_name)
    # 维护一些样式， style_heading, style_body, style_red, style_green
    style_heading, style_body = excel_download_style()

    # 写标题栏
    for index, head_name in enumerate(head_list):
        sheet.write(0, index, head_name, style_heading)

    # 写数据
    row = 1
    for query_obj in queryset:
        query_dict = model_to_dict(query_obj)
        for index, head_name in enumerate(head_list):
            sheet.write(row, index, query_dict.get(head_name, None), style_body)
        row = row + 1
    wb.save(filename)


def excel_download_style():
    # 维护一些样式， style_heading, style_body, style_red, style_green
    style_heading = xlwt.easyxf("""
            font:
                name Arial,
                colour_index white,
                bold on,
                height 0xA0;
            align:
                wrap off,
                vert center,
                horiz center;
            pattern:
                pattern solid,
                fore-colour 0x19;
            borders:
                left THIN,
                right THIN,
                top THIN,
                bottom THIN;
            """
                                )
    style_body = xlwt.easyxf("""
                font:
                    name Arial,
                    bold off,
                    height 0XA0;
                align:
                    wrap on,
                    vert center,
                    horiz left;
                borders:
                    left THIN,
                    right THIN,
                    top THIN,
                    bottom THIN;
                """
                             )
    # style_green = xlwt.easyxf(" pattern: pattern solid,fore-colour 0x11;")
    # style_red = xlwt.easyxf(" pattern: pattern solid,fore-colour 0x0A;")
    fmts = [
        'M/D/YY',
        'D-MMM-YY',
        'D-MMM',
        'MMM-YY',
        'h:mm AM/PM',
        'h:mm:ss AM/PM',
        'h:mm',
        'h:mm:ss',
        'M/D/YY h:mm',
        'mm:ss',
        '[h]:mm:ss',
        'mm:ss.0',
    ]
    style_body.num_format_str = fmts[0]
    return style_heading, style_body
